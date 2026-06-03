#!/usr/bin/env python
"""Ingest the official NOMENCLATURE-*.xlsx into ``fixtures/nomenclature.csv``.

The real workbook (``data.yaml: nomenclature.source_xlsx``) is a formatted ledger:
a multi-row Republique-Algerienne banner, then a header row, then thousands of
drug rows. Its exact column titles are not part of any contract, so this script:

  1. scans the first rows to locate the **header row** (the row carrying the most
     known column tokens), unless ``--header-row`` is given;
  2. maps each real header to our schema via the documented :data:`COLUMN_MAP`
     below (best-effort, token-based, accent/encoding tolerant);
  3. streams the data rows out as the CSV columns configured in
     ``configs/nomenclature/correction.yaml`` (``csv.columns`` / ``csv.encoding``).

The loader/matcher own *consuming* this CSV; this script only *produces* it.

TODO(real-data): CONFIRM THE MAPPING below against the shipped workbook before
relying on the output. The header tokens were read from
``NOMENCLATURE-VERSION-FEVRIER-2026-.xlsx`` (sheet "Nomenclature Fevrier 2026",
header at row 14, 1-based) and are encoding-tolerant, but:
  * ``num_enregistrement`` in the real file looks like ``352/01 A 003/06/22`` —
    a DIFFERENT shape from the synthetic ``AA/BB/CC<LETTER>DDD/EEE``. Reconcile the
    canonical code format (parsing/fields.yaml) with real data before switchover.
  * ``tr`` (taux de remboursement) has NO obvious source column in the observed
    header; it is left blank here. Map it once the correct column is identified.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

# Make the package importable when run as a loose script (python scripts/...).
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from vignocr.common import get_logger, load_config  # noqa: E402
from vignocr.common.config import resolve_path  # noqa: E402

log = get_logger(__name__)

# --------------------------------------------------------------------------- #
# COLUMN_MAP — our schema column -> list of accepted header tokens (normalized:
# accent/case/space-insensitive). The FIRST header cell that CONTAINS any token
# wins. Order tokens most-specific-first. See module TODO(real-data).
# --------------------------------------------------------------------------- #
COLUMN_MAP: dict[str, list[str]] = {
    # "N°ENREGISTREMENT" (avoid bare "n" which would also hit "N°" / "nom").
    "num_enregistrement": ["nenregistrement", "n enregistrement", "enregistrement"],
    # "NOM DE MARQUE" is the commercial product name.
    "product_name": ["nom de marque", "marque", "nom commercial"],
    # "DENOMINATION COMMUNE INTERNATIONALE" (DCI).
    "dci": ["denomination commune", "dci", "denomination"],
    "dosage": ["dosage", "dose"],
    "forme": ["forme"],
    # "LABORATOIRES DETENTEUR DE LA DECISION ..."
    "laboratoire": ["laboratoire", "laboratoires", "fabricant", "detenteur"],
    # TODO(real-data): no observed source column for the reimbursement rate in the
    # Feb-2026 file (it carries PAYS/LISTE/P1/P2/OBS, none of which is the taux).
    # Tokens are kept SPECIFIC so we never mis-claim a lookalike header (a bare
    # "tr" matches inside "enregis-TR-ement"); left unmapped -> emitted blank.
    "tr": ["taux de remboursement", "taux remboursement", "remboursement"],
}

# Header detection: a row is "the header" if it contains at least this many of the
# expected tokens across all schema columns.
_MIN_HEADER_HITS = 4
_HEADER_SCAN_ROWS = 40


def _norm(text: Any) -> str:
    """Accent/case/space-insensitive normalization for header matching."""
    if text is None:
        return ""
    s = str(text)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("°", "").replace("º", "")
    s = re.sub(r"[^a-z0-9 ]+", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def _row_header_hits(cells: tuple[Any, ...]) -> int:
    """How many schema columns find a token match in this row's cells."""
    normed = [_norm(c) for c in cells]
    hits = 0
    for tokens in COLUMN_MAP.values():
        if any(any(tok in cell for cell in normed) for tok in tokens):
            hits += 1
    return hits


def detect_header_row(ws: Any, scan_rows: int = _HEADER_SCAN_ROWS) -> int:
    """Return the 1-based index of the best header row, or raise if none qualifies."""
    best_row, best_hits = -1, 0
    for i, cells in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1
    ):
        hits = _row_header_hits(cells)
        if hits > best_hits:
            best_row, best_hits = i, hits
    if best_row < 0 or best_hits < _MIN_HEADER_HITS:
        raise ValueError(
            f"could not locate a header row in the first {scan_rows} rows "
            f"(best match had {best_hits} known columns; need >= {_MIN_HEADER_HITS}). "
            f"Pass --header-row to set it explicitly."
        )
    return best_row


def build_column_index(header_cells: tuple[Any, ...]) -> dict[str, int]:
    """Map each schema column to a 0-based source column index (best-effort).

    A header cell is claimed by the first schema column whose tokens it contains;
    each source cell is used at most once. Schema columns with no match are absent.
    """
    normed = [_norm(c) for c in header_cells]
    claimed: set[int] = set()
    mapping: dict[str, int] = {}
    for schema_col, tokens in COLUMN_MAP.items():
        for tok in tokens:
            idx = next(
                (j for j, cell in enumerate(normed) if j not in claimed and tok and tok in cell),
                None,
            )
            if idx is not None:
                mapping[schema_col] = idx
                claimed.add(idx)
                break
    return mapping


def _cell_to_str(value: Any) -> str:
    """Render a cell value as a clean string (dates -> ISO date, else str)."""
    if value is None:
        return ""
    if hasattr(value, "date") and hasattr(value, "year"):  # datetime/date
        try:
            return value.date().isoformat() if hasattr(value, "date") else value.isoformat()
        except Exception:  # noqa: BLE001 - never let a stray cell type abort ingest
            return str(value).strip()
    return str(value).strip()


def ingest(
    xlsx_path: Path,
    out_path: Path,
    *,
    sheet: str | None = None,
    header_row: int | None = None,
    limit: int | None = None,
    cfg: dict[str, Any] | None = None,
) -> int:
    """Convert the xlsx to the configured CSV. Returns the number of data rows written."""
    cfg = cfg if cfg is not None else load_config("nomenclature/correction")
    csv_cfg = cfg.get("csv", {})
    columns: list[str] = list(csv_cfg.get("columns", []))
    encoding: str = csv_cfg.get("encoding", "utf-8")

    if not xlsx_path.exists():
        raise FileNotFoundError(f"source xlsx not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        log.info("ingest.open", xlsx=str(xlsx_path), sheet=ws.title)

        hdr_idx = header_row if header_row is not None else detect_header_row(ws)
        header_cells = next(ws.iter_rows(min_row=hdr_idx, max_row=hdr_idx, values_only=True))
        col_index = build_column_index(header_cells)

        mapped = {c: header_cells[i] for c, i in col_index.items()}
        log.info("ingest.header", row=hdr_idx, mapping=mapped)
        missing = [c for c in columns if c not in col_index]
        if missing:
            log.warning("ingest.unmapped_columns", columns=missing)

        out_path.parent.mkdir(parents=True, exist_ok=True)
        written = 0
        with open(out_path, "w", encoding=encoding, newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns)
            writer.writeheader()
            for cells in ws.iter_rows(min_row=hdr_idx + 1, values_only=True):
                record = {}
                for col in columns:
                    src = col_index.get(col)
                    record[col] = (
                        _cell_to_str(cells[src]) if src is not None and src < len(cells) else ""
                    )
                # Skip fully-blank rows (trailing sheet padding, separators).
                if not any(record.values()):
                    continue
                writer.writerow(record)
                written += 1
                if limit is not None and written >= limit:
                    break
    finally:
        wb.close()

    log.info("ingest.done", out=str(out_path), rows=written, columns=columns)
    return written


def build_parser() -> argparse.ArgumentParser:
    cfg = load_config("nomenclature/correction")
    default_out = cfg.get("csv", {}).get("path", "fixtures/nomenclature.csv")
    # Default source from data.yaml (single source of truth for the xlsx name).
    data_cfg = load_config("data").get("nomenclature", {})
    default_xlsx = data_cfg.get("source_xlsx", "NOMENCLATURE-VERSION-FEVRIER-2026-.xlsx")

    p = argparse.ArgumentParser(
        description="Convert the official NOMENCLATURE xlsx into the fixtures CSV.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--xlsx", default=default_xlsx, help="Source .xlsx (repo-root-relative ok).")
    p.add_argument("--out", default=default_out, help="Destination CSV (repo-root-relative ok).")
    p.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet).")
    p.add_argument(
        "--header-row",
        type=int,
        default=None,
        help="1-based header row index (default: auto-detect).",
    )
    p.add_argument("--limit", type=int, default=None, help="Max data rows to write (default: all).")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    xlsx_path = resolve_path(args.xlsx)
    out_path = resolve_path(args.out)
    try:
        n = ingest(
            xlsx_path,
            out_path,
            sheet=args.sheet,
            header_row=args.header_row,
            limit=args.limit,
        )
    except (FileNotFoundError, ValueError) as exc:
        log.error("ingest.failed", error=str(exc))
        print(f"error: {exc}", file=sys.stderr)
        return 1
    print(f"wrote {n} rows -> {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
